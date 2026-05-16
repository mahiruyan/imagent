from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models.company import Company


EXPORT_COLUMNS = [
    "Supplier ID",
    "Company Name",
    "City",
    "District / OSB",
    "Category",
    "Address",
    "Website",
    "Phone",
    "Email",
    "Rating",
    "Review Count",
    "Source Providers",
    "Notes",
    "Status",
    "Added Date",
]


class ExportService:
    def build_workbook_bytes(self, companies: list[Company]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Suppliers"
        sheet.append(EXPORT_COLUMNS)

        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        for company in companies:
            sheet.append(
                [
                    company.supplier_code,
                    company.canonical_name,
                    company.city,
                    company.osb or company.district,
                    "",
                    company.canonical_address,
                    company.website,
                    company.phone,
                    company.email,
                    float(company.rating) if company.rating is not None else None,
                    company.review_count,
                    "",
                    company.notes,
                    company.status,
                    company.created_at.date().isoformat() if company.created_at else None,
                ]
            )

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

